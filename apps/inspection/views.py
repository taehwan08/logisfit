"""
검수 시스템 뷰
"""
import re
import json
from collections import defaultdict

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.utils import timezone
import openpyxl
import xlrd

from .models import Order, OrderProduct, InspectionLog, UploadBatch


# ============================================================================
# 엑셀 파싱 유틸리티
# ============================================================================

def _parse_excel(excel_file):
    """엑셀 파일을 파싱하여 헤더와 데이터 행을 반환한다. xlsx/xls 모두 지원."""
    filename = excel_file.name.lower()

    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    elif filename.endswith('.xls'):
        content = excel_file.read()
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        headers = [ws.cell_value(0, col) for col in range(ws.ncols)]
        rows = []
        for row_idx in range(1, ws.nrows):
            rows.append(tuple(ws.cell_value(row_idx, col) for col in range(ws.ncols)))
    else:
        raise ValueError('엑셀 파일만 업로드 가능합니다.')

    return headers, rows


def _get_col(row, col_map, key, default=''):
    """행에서 컬럼 값을 안전하게 가져온다."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None:
        return default
    return str(val).strip()


def _detect_format(headers):
    """헤더를 보고 양식1 / 양식2를 자동 판별한다.

    양식1 식별: '쇼핑몰', '바코드번호', '매칭상품명' 컬럼 존재
    양식2 식별: '상품명' 컬럼 존재 + '수취인명(받는분)' 컬럼 존재

    Returns: 'format1' | 'format2'
    """
    header_set = set(headers)
    if {'쇼핑몰', '바코드번호', '매칭상품명'}.issubset(header_set):
        return 'format1'
    if '상품명' in header_set and any('수취인' in h or '받는분' in h for h in headers):
        return 'format2'
    # fallback - 양식1 필수 컬럼으로 재시도
    if '송장번호' in header_set and '매칭수량' in header_set:
        return 'format1'
    return 'unknown'


def _parse_format2_product_cell(product_text):
    """양식2 상품명 셀을 파싱하여 상품 리스트를 반환한다.

    입력 예시:
    "[이옴] 트러블 컨트롤 패드 140ml(70p)[90019]●1개 [P1-8800309590019]|[이옴] 트러블 패치 마스크 20ml(4p)[90057]●1개 [P1-8800309590057]"

    파싱 규칙:
    - "|"로 상품 구분
    - "●" 뒤에 숫자 → 수량
    - "[P1-바코드]" 또는 유사 패턴에서 "-" 뒤가 바코드번호
    - "●" 앞이 상품명

    Returns: [{'product_name': ..., 'barcode': ..., 'quantity': ...}, ...]
    """
    if not product_text or not product_text.strip():
        return []

    items = product_text.split('|')
    products = []

    for item in items:
        item = item.strip()
        if not item:
            continue

        # 바코드 추출: [P1-8800309590019] 또는 [XX-바코드] 패턴
        barcode = ''
        barcode_match = re.search(r'\[[\w]+-(\d+)\]', item)
        if barcode_match:
            barcode = barcode_match.group(1)

        # 수량 추출: ●1개, ●2개 등
        quantity = 1
        qty_match = re.search(r'●\s*(\d+)\s*개', item)
        if qty_match:
            quantity = int(qty_match.group(1))

        # 상품명 추출: ● 앞부분
        product_name = item.split('●')[0].strip() if '●' in item else item.strip()
        # 상품명에서 후행 바코드 태그 제거
        product_name = re.sub(r'\s*\[[\w]+-\d+\]\s*$', '', product_name).strip()

        if barcode and product_name:
            products.append({
                'product_name': product_name,
                'barcode': barcode,
                'quantity': quantity,
            })

    return products


def _process_format1(headers, rows):
    """양식1 (쇼핑몰/바코드번호/매칭상품명 등 컬럼이 개별 존재) 처리.

    Returns: (orders_data dict, batch_print_order, batch_delivery_memo, error_message or None)
    """
    COLUMN_MAP = {
        '송장번호': '송장번호',
        '쇼핑몰': '판매처',
        '수령자': '수령인',
        '전화1': '핸드폰',
        '주소': '주소',
        '바코드번호': '상품바코드',
        '매칭상품명': '매칭상품명',
        '매칭관리명': '매칭관리명',
        '매칭수량': '수량',
        '출력차수': '출력차수',
        '배송메모': '배송메모',
        '등록일': '등록일',
        '택배사': '택배사',
    }

    required_excel_columns = ['송장번호', '쇼핑몰', '수령자', '전화1', '주소', '바코드번호', '매칭상품명', '매칭수량']

    col_map = {}
    for excel_col, internal_key in COLUMN_MAP.items():
        if excel_col in headers:
            col_map[internal_key] = headers.index(excel_col)

    missing = []
    for excel_col in required_excel_columns:
        internal_key = COLUMN_MAP[excel_col]
        if internal_key not in col_map:
            missing.append(excel_col)
    if missing:
        return None, '', '', f'필수 컬럼이 없습니다: {", ".join(missing)}'

    orders_data = defaultdict(lambda: {'info': None, 'products': []})
    batch_print_order = ''
    batch_delivery_memo = ''

    for row in rows:
        tracking_number = _get_col(row, col_map, '송장번호')
        if not tracking_number:
            continue

        if not batch_print_order:
            batch_print_order = _get_col(row, col_map, '출력차수')
        if not batch_delivery_memo:
            batch_delivery_memo = _get_col(row, col_map, '배송메모')

        if orders_data[tracking_number]['info'] is None:
            orders_data[tracking_number]['info'] = {
                'seller': _get_col(row, col_map, '판매처'),
                'receiver_name': _get_col(row, col_map, '수령인'),
                'receiver_phone': _get_col(row, col_map, '핸드폰'),
                'receiver_address': _get_col(row, col_map, '주소'),
                'registered_date': _get_col(row, col_map, '등록일'),
                'courier': _get_col(row, col_map, '택배사'),
                'print_order': _get_col(row, col_map, '출력차수'),
                'delivery_memo': _get_col(row, col_map, '배송메모'),
            }

        barcode = _get_col(row, col_map, '상품바코드')
        product_name_part = _get_col(row, col_map, '매칭상품명')
        manage_name_part = _get_col(row, col_map, '매칭관리명')

        if product_name_part and manage_name_part:
            product_name = f'{product_name_part} ({manage_name_part})'
        elif product_name_part:
            product_name = product_name_part
        else:
            product_name = manage_name_part

        try:
            qty_val = row[col_map['수량']] if col_map.get('수량') is not None else 0
            quantity = int(float(str(qty_val or 0)))
        except (ValueError, TypeError, IndexError):
            quantity = 0

        if barcode and product_name:
            orders_data[tracking_number]['products'].append({
                'barcode': barcode,
                'product_name': product_name,
                'quantity': quantity,
            })

    return orders_data, batch_print_order, batch_delivery_memo, None


def _process_format2(headers, rows):
    """양식2 (상품명 셀에 바코드+수량이 합쳐진 형태) 처리.

    필요 컬럼: 송장번호, 상품명, 수취인명(받는분), 수취인명(받는분) 핸드폰번호, 수취인명(받는분) 주소1

    Returns: (orders_data dict, batch_print_order, batch_delivery_memo, error_message or None)
    """
    # 유연한 헤더 매칭 (부분 일치)
    col_map = {}
    for idx, h in enumerate(headers):
        hl = h.lower().strip() if h else ''
        if h == '송장번호':
            col_map['송장번호'] = idx
        elif h == '상품명':
            col_map['상품명'] = idx
        elif '핸드폰' in h or '전화' in h or '연락처' in h:
            col_map['핸드폰'] = idx
        elif '주소' in h:
            # 주소1 우선, 아니면 일반 주소
            if '주소1' in h or '주소' == h.strip():
                col_map['주소'] = idx
            elif '주소' not in col_map:
                col_map['주소'] = idx
        elif '수취인' in h or '받는분' in h:
            # 핸드폰/주소가 아닌 수취인명 컬럼
            if '핸드폰' not in h and '전화' not in h and '주소' not in h:
                col_map['수령인'] = idx

    # 필수 컬럼 체크
    required = {'송장번호': '송장번호', '상품명': '상품명', '수령인': '수취인명(받는분)'}
    missing = []
    for key, display in required.items():
        if key not in col_map:
            missing.append(display)
    if missing:
        return None, '', '', f'필수 컬럼이 없습니다: {", ".join(missing)}'

    orders_data = defaultdict(lambda: {'info': None, 'products': []})

    for row in rows:
        tracking_number = _get_col(row, col_map, '송장번호')
        if not tracking_number:
            continue

        if orders_data[tracking_number]['info'] is None:
            orders_data[tracking_number]['info'] = {
                'seller': '',
                'receiver_name': _get_col(row, col_map, '수령인'),
                'receiver_phone': _get_col(row, col_map, '핸드폰'),
                'receiver_address': _get_col(row, col_map, '주소'),
                'registered_date': '',
                'courier': '',
                'print_order': '',
                'delivery_memo': '',
            }

        product_text = _get_col(row, col_map, '상품명')
        parsed_products = _parse_format2_product_cell(product_text)

        for p in parsed_products:
            # 중복 상품 방지 (같은 송장에 같은 바코드가 이미 있으면 스킵)
            existing = [ep for ep in orders_data[tracking_number]['products'] if ep['barcode'] == p['barcode']]
            if not existing:
                orders_data[tracking_number]['products'].append(p)

    return orders_data, '', '', None


def office_page(request):
    """오피스팀 페이지"""
    orders = Order.objects.all()
    total = orders.count()
    waiting = orders.filter(status='대기중').count()
    inspecting = orders.filter(status='검수중').count()
    completed = orders.filter(status='완료').count()
    context = {
        'total': total,
        'waiting': waiting,
        'inspecting': inspecting,
        'completed': completed,
    }
    return render(request, 'inspection/office.html', context)


def field_page(request):
    """필드팀 페이지"""
    return render(request, 'inspection/field.html')


@csrf_exempt
@require_POST
def upload_excel(request):
    """엑셀 업로드 처리 - 양식1(쇼핑몰/바코드번호 개별컬럼), 양식2(상품명 합산셀) 자동 감지"""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': '파일이 없습니다.'}, status=400)

    excel_file = request.FILES['file']

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': '엑셀 파일만 업로드 가능합니다.'}, status=400)

    try:
        headers, rows = _parse_excel(excel_file)
        headers = [str(h).strip() if h else '' for h in headers]

        # 양식 자동 감지
        file_format = _detect_format(headers)

        if file_format == 'format1':
            orders_data, batch_print_order, batch_delivery_memo, error = _process_format1(headers, rows)
        elif file_format == 'format2':
            orders_data, batch_print_order, batch_delivery_memo, error = _process_format2(headers, rows)
        else:
            return JsonResponse({
                'success': False,
                'message': '인식할 수 없는 엑셀 양식입니다. 양식1(쇼핑몰/바코드번호) 또는 양식2(상품명 합산) 형식이 필요합니다.'
            }, status=400)

        if error:
            return JsonResponse({'success': False, 'message': error}, status=400)

        if not orders_data:
            return JsonResponse({'success': False, 'message': '유효한 데이터가 없습니다.'}, status=400)

        # DB 저장
        total_orders = 0
        total_products = 0
        duplicated = 0

        with transaction.atomic():
            uploader = ''
            if request.user.is_authenticated:
                uploader = request.user.name or request.user.email or ''

            batch = UploadBatch.objects.create(
                file_name=excel_file.name,
                print_order=batch_print_order,
                delivery_memo=batch_delivery_memo,
                uploaded_by=uploader,
            )

            for tracking_number, data in orders_data.items():
                existing = Order.objects.filter(tracking_number=tracking_number)
                if existing.exists():
                    existing.delete()
                    duplicated += 1

                order = Order.objects.create(
                    upload_batch=batch,
                    tracking_number=tracking_number,
                    **data['info']
                )
                total_orders += 1

                products = [
                    OrderProduct(
                        order=order,
                        barcode=p['barcode'],
                        product_name=p['product_name'],
                        quantity=p['quantity'],
                    )
                    for p in data['products']
                ]
                OrderProduct.objects.bulk_create(products)
                total_products += len(products)

            batch.total_orders = total_orders
            batch.total_products = total_products
            batch.save(update_fields=['total_orders', 'total_products'])

        format_label = '양식1' if file_format == 'format1' else '양식2'
        message = f'{total_orders}건의 송장이 등록되었습니다. ({format_label})'
        if duplicated:
            message += f' (중복 {duplicated}건 재등록)'

        return JsonResponse({
            'success': True,
            'message': message,
            'total_orders': total_orders,
            'total_products': total_products,
            'duplicated': duplicated,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'파일 처리 중 오류: {str(e)}'}, status=500)


@require_GET
def get_order(request, tracking_number):
    """송장 조회"""
    worker = request.user.name if request.user.is_authenticated else None

    try:
        order = Order.objects.prefetch_related('products').get(
            tracking_number=tracking_number
        )

        # 기처리 배송 체크
        if order.status == '완료':
            InspectionLog.objects.create(
                tracking_number=tracking_number,
                scan_type='송장',
                alert_code='기처리배송',
                worker=worker,
            )
            return JsonResponse({
                'success': True,
                'alert_code': '기처리배송',
                'order': {
                    'tracking_number': order.tracking_number,
                    'seller': order.seller,
                    'receiver_name': order.receiver_name,
                    'receiver_phone': order.receiver_phone,
                    'receiver_address': order.receiver_address,
                    'status': order.status,
                    'completed_at': order.completed_at.isoformat() if order.completed_at else None,
                },
                'products': [
                    {
                        'id': p.id,
                        'barcode': p.barcode,
                        'product_name': p.product_name,
                        'quantity': p.quantity,
                        'scanned_quantity': p.scanned_quantity,
                    }
                    for p in order.products.all()
                ],
            })

        # 정상 조회 로그
        InspectionLog.objects.create(
            tracking_number=tracking_number,
            scan_type='송장',
            alert_code='정상',
            worker=worker,
        )

        return JsonResponse({
            'success': True,
            'alert_code': '정상',
            'order': {
                'tracking_number': order.tracking_number,
                'seller': order.seller,
                'receiver_name': order.receiver_name,
                'receiver_phone': order.receiver_phone,
                'receiver_address': order.receiver_address,
                'status': order.status,
            },
            'products': [
                {
                    'id': p.id,
                    'barcode': p.barcode,
                    'product_name': p.product_name,
                    'quantity': p.quantity,
                    'scanned_quantity': p.scanned_quantity,
                }
                for p in order.products.all()
            ],
        })

    except Order.DoesNotExist:
        InspectionLog.objects.create(
            tracking_number=tracking_number,
            scan_type='송장',
            alert_code='송장번호미등록',
            worker=worker,
        )
        return JsonResponse({
            'success': False,
            'alert_code': '송장번호미등록',
            'message': '등록되지 않은 송장번호입니다.',
        })


@csrf_exempt
@require_POST
def scan_product(request):
    """상품 스캔 처리"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'}, status=400)

    tracking_number = data.get('tracking_number', '').strip()
    barcode = data.get('barcode', '').strip()
    worker = request.user.name if request.user.is_authenticated else None

    if not tracking_number or not barcode:
        return JsonResponse({'success': False, 'message': '송장번호와 바코드가 필요합니다.'}, status=400)

    try:
        with transaction.atomic():
            order = Order.objects.select_for_update().get(tracking_number=tracking_number)

            # 해당 주문의 상품 중 바코드 매칭
            products = list(order.products.select_for_update().filter(barcode=barcode))

            if not products:
                # 상품오류
                InspectionLog.objects.create(
                    tracking_number=tracking_number,
                    barcode=barcode,
                    scan_type='상품',
                    alert_code='상품오류',
                    worker=worker,
                )
                return JsonResponse({
                    'success': False,
                    'alert_code': '상품오류',
                    'message': '해당 송장에 없는 상품입니다.',
                })

            # 스캔 가능한 상품 찾기 (scanned_quantity < quantity)
            target_product = None
            for p in products:
                if p.scanned_quantity < p.quantity:
                    target_product = p
                    break

            if target_product is None:
                # 중복스캔
                InspectionLog.objects.create(
                    tracking_number=tracking_number,
                    barcode=barcode,
                    scan_type='상품',
                    alert_code='중복스캔',
                    worker=worker,
                )
                return JsonResponse({
                    'success': False,
                    'alert_code': '중복스캔',
                    'message': '이미 스캔 완료된 상품입니다.',
                })

            # 스캔 수량 증가
            target_product.scanned_quantity += 1
            target_product.save(update_fields=['scanned_quantity'])

            # Order status 업데이트
            if order.status == '대기중':
                order.status = '검수중'
                order.save(update_fields=['status'])

            # 남은 수량 계산 (해당 상품)
            remaining = target_product.quantity - target_product.scanned_quantity

            # 전체 완료 체크
            all_products = list(order.products.all())
            all_completed = all(p.scanned_quantity >= p.quantity for p in all_products)

            if all_completed:
                order.status = '완료'
                order.completed_at = timezone.now()
                order.save(update_fields=['status', 'completed_at'])

                InspectionLog.objects.create(
                    tracking_number=tracking_number,
                    barcode=barcode,
                    scan_type='상품',
                    alert_code='완료',
                    worker=worker,
                )

                return JsonResponse({
                    'success': True,
                    'alert_code': '완료',
                    'all_completed': True,
                    'product': {
                        'id': target_product.id,
                        'barcode': target_product.barcode,
                        'product_name': target_product.product_name,
                        'quantity': target_product.quantity,
                        'scanned_quantity': target_product.scanned_quantity,
                    },
                    'order': {
                        'tracking_number': order.tracking_number,
                        'completed_at': order.completed_at.isoformat(),
                    },
                })

            # 남은 수량에 따른 알림코드
            if remaining > 0:
                alert_code = '숫자'
            else:
                alert_code = '정상'

            InspectionLog.objects.create(
                tracking_number=tracking_number,
                barcode=barcode,
                scan_type='상품',
                alert_code=alert_code,
                worker=worker,
            )

            return JsonResponse({
                'success': True,
                'alert_code': alert_code,
                'remaining': remaining,
                'product': {
                    'id': target_product.id,
                    'barcode': target_product.barcode,
                    'product_name': target_product.product_name,
                    'quantity': target_product.quantity,
                    'scanned_quantity': target_product.scanned_quantity,
                },
                'all_completed': False,
                'products': [
                    {
                        'id': p.id,
                        'barcode': p.barcode,
                        'product_name': p.product_name,
                        'quantity': p.quantity,
                        'scanned_quantity': p.scanned_quantity,
                    }
                    for p in all_products
                ],
            })

    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'alert_code': '송장번호미등록',
            'message': '등록되지 않은 송장번호입니다.',
        })


@csrf_exempt
@require_POST
def complete_inspection(request):
    """검수 완료 처리"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'}, status=400)

    tracking_number = data.get('tracking_number', '').strip()
    worker = request.user.name if request.user.is_authenticated else None

    if not tracking_number:
        return JsonResponse({'success': False, 'message': '송장번호가 필요합니다.'}, status=400)

    try:
        with transaction.atomic():
            order = Order.objects.select_for_update().get(tracking_number=tracking_number)

            # 모든 상품 스캔 완료 여부 확인
            all_products = order.products.all()
            incomplete = [p for p in all_products if p.scanned_quantity < p.quantity]

            if incomplete:
                return JsonResponse({
                    'success': False,
                    'message': f'아직 스캔하지 않은 상품이 {len(incomplete)}건 있습니다.',
                })

            order.status = '완료'
            order.completed_at = timezone.now()
            order.save(update_fields=['status', 'completed_at'])

            InspectionLog.objects.create(
                tracking_number=tracking_number,
                scan_type='상품',
                alert_code='완료',
                worker=worker,
            )

            return JsonResponse({
                'success': True,
                'message': '검수가 완료되었습니다.',
                'tracking_number': tracking_number,
                'completed_at': order.completed_at.isoformat(),
            })

    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '등록되지 않은 송장번호입니다.',
        })


@require_GET
def get_logs(request):
    """로그 조회"""
    logs = InspectionLog.objects.all()

    tracking_number = request.GET.get('tracking_number')
    if tracking_number:
        logs = logs.filter(tracking_number=tracking_number)

    date_from = request.GET.get('date_from')
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    alert_code = request.GET.get('alert_code')
    if alert_code:
        logs = logs.filter(alert_code=alert_code)

    total = logs.count()
    logs = logs[:200]

    return JsonResponse({
        'success': True,
        'logs': [
            {
                'id': log.id,
                'tracking_number': log.tracking_number,
                'barcode': log.barcode,
                'scan_type': log.scan_type,
                'alert_code': log.alert_code,
                'worker': log.worker,
                'created_at': log.created_at.isoformat(),
            }
            for log in logs
        ],
        'total': total,
    })


@require_GET
def get_upload_batches(request):
    """업로드 이력 조회 (오피스팀용)"""
    batches = UploadBatch.objects.all()

    total = batches.count()
    batches = batches[:100]

    result = []
    for b in batches:
        orders = b.orders.all()
        order_count = orders.count()
        waiting = orders.filter(status='대기중').count()
        inspecting = orders.filter(status='검수중').count()
        completed = orders.filter(status='완료').count()

        result.append({
            'id': b.id,
            'file_name': b.file_name,
            'print_order': b.print_order,
            'delivery_memo': b.delivery_memo,
            'total_orders': b.total_orders,
            'total_products': b.total_products,
            'waiting': waiting,
            'inspecting': inspecting,
            'completed': completed,
            'uploaded_by': b.uploaded_by,
            'uploaded_at': b.uploaded_at.isoformat(),
        })

    return JsonResponse({
        'success': True,
        'batches': result,
        'total': total,
    })


@csrf_exempt
@require_POST
def delete_upload_batch(request, batch_id):
    """업로드 배치 삭제 (해당 배치의 모든 송장/상품 함께 삭제)"""
    try:
        batch = UploadBatch.objects.get(id=batch_id)
        file_name = batch.file_name
        batch.delete()  # CASCADE로 orders, products 모두 삭제
        return JsonResponse({
            'success': True,
            'message': f'"{file_name}" 업로드 데이터가 삭제되었습니다.',
        })
    except UploadBatch.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '해당 업로드 이력을 찾을 수 없습니다.',
        }, status=404)
