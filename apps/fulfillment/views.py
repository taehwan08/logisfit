"""
출고 관리 뷰

B2B 출고 주문 목록, 등록, 수정, 삭제, 상태변경, 벌크 붙여넣기, 엑셀 다운로드 기능을 제공합니다.
"""
import json
import logging
import os
import re
from datetime import datetime
from functools import wraps

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone

from .models import FulfillmentOrder, FulfillmentComment, PlatformColumnConfig
from .slack import send_order_created_notification, send_bulk_orders_notification
from apps.accounts.email import send_shipment_notification_async, send_shipment_notifications_async
from apps.clients.models import Client, Brand

logger = logging.getLogger(__name__)

# 댓글 첨부파일 설정
COMMENT_ALLOWED_EXTENSIONS = {
    '.jpg', '.jpeg', '.png', '.gif', '.webp',  # 이미지
    '.pdf', '.xlsx', '.xls', '.doc', '.docx',  # 문서
}
COMMENT_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
COMMENT_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# 고정 모델 필드 (platform_data가 아닌 모델에 직접 저장)
FIXED_MODEL_FIELDS = {'product_name', 'quantity', 'box_quantity', 'pallet_quantity', 'invoice_number'}


# ============================================================================
# 권한 데코레이터
# ============================================================================

def fulfillment_access_required(view_func):
    """출고 관리 접근 가능 데코레이터 (관리자/작업자/고객사)"""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        if not (user.is_admin or user.is_worker or user.is_client or user.is_superuser):
            return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_or_worker_required(view_func):
    """관리자 또는 작업자 전용 데코레이터"""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        if not (user.is_admin or user.is_worker or user.is_superuser):
            return JsonResponse({'error': '관리자/작업자 권한이 필요합니다.'}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_required(view_func):
    """관리자 전용 데코레이터"""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        user = request.user
        if not (user.is_admin or user.is_superuser):
            return JsonResponse({'error': '관리자 권한이 필요합니다.'}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_client_filter(user):
    """고객사 사용자의 경우 자기 거래처만 필터링"""
    if user.is_client:
        client_ids = user.clients.values_list('id', flat=True)
        return Q(client_id__in=client_ids)
    return Q()  # 관리자/작업자는 전체


# ============================================================================
# 페이지 뷰
# ============================================================================

@fulfillment_access_required
def order_list_page(request):
    """출고 주문 목록 페이지"""
    user = request.user

    # 고객사 목록 (필터용)
    if user.is_client:
        clients = user.clients.filter(is_active=True)
    else:
        clients = Client.objects.filter(is_active=True).order_by('company_name')

    context = {
        'clients': clients,
        'platforms': FulfillmentOrder.Platform.choices,
        'statuses': FulfillmentOrder.Status.choices,
        'is_admin': user.is_admin or user.is_superuser,
        'is_client': user.is_client,
        'is_worker': user.is_worker,
    }
    return render(request, 'fulfillment/order_list.html', context)


# ============================================================================
# 브랜드 API
# ============================================================================

@fulfillment_access_required
@require_http_methods(["GET"])
def get_brands(request):
    """거래처별 브랜드 목록 조회"""
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'brands': []})

    # 거래처 사용자는 자기 거래처의 브랜드만 조회 가능
    user = request.user
    if user.is_client and not user.clients.filter(id=client_id).exists():
        return JsonResponse({'brands': []})

    brands = Brand.objects.filter(
        client_id=client_id,
        is_active=True,
    ).order_by('name')

    brand_list = [{'id': b.id, 'name': b.name, 'code': b.code} for b in brands]
    return JsonResponse({'brands': brand_list})


@admin_or_worker_required
@require_http_methods(["POST"])
def create_brand(request):
    """브랜드 등록 (관리자/작업자)"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    client_id = data.get('client_id')
    name = data.get('name', '').strip()

    if not client_id or not name:
        return JsonResponse({'error': '거래처와 브랜드명을 입력해주세요.'}, status=400)

    try:
        client = Client.objects.get(id=client_id, is_active=True)
    except Client.DoesNotExist:
        return JsonResponse({'error': '유효하지 않은 거래처입니다.'}, status=400)

    if Brand.objects.filter(client=client, name=name).exists():
        return JsonResponse({'error': '이미 등록된 브랜드명입니다.'}, status=400)

    brand = Brand.objects.create(
        client=client,
        name=name,
        code=data.get('code', '').strip(),
        memo=data.get('memo', '').strip(),
        created_by=request.user,
    )

    return JsonResponse({
        'success': True,
        'message': '브랜드가 등록되었습니다.',
        'brand': {'id': brand.id, 'name': brand.name, 'code': brand.code},
    })


@admin_or_worker_required
@require_http_methods(["POST"])
def update_brand(request, brand_id):
    """브랜드 수정 (관리자/작업자)"""
    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'error': '브랜드를 찾을 수 없습니다.'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    if 'name' in data:
        name = data['name'].strip()
        if name:
            # 중복 체크 (같은 거래처 내)
            dup = Brand.objects.filter(client=brand.client, name=name).exclude(id=brand.id)
            if dup.exists():
                return JsonResponse({'error': '이미 등록된 브랜드명입니다.'}, status=400)
            brand.name = name
    if 'code' in data:
        brand.code = data['code'].strip()
    if 'memo' in data:
        brand.memo = data['memo'].strip()
    if 'is_active' in data:
        brand.is_active = data['is_active']

    brand.save()
    return JsonResponse({'success': True, 'message': '브랜드가 수정되었습니다.'})


@admin_or_worker_required
@require_http_methods(["POST"])
def delete_brand(request, brand_id):
    """브랜드 삭제 (관리자/작업자)"""
    try:
        brand = Brand.objects.get(id=brand_id)
    except Brand.DoesNotExist:
        return JsonResponse({'error': '브랜드를 찾을 수 없습니다.'}, status=404)

    # 해당 브랜드에 연결된 주문이 있으면 비활성화만
    if brand.fulfillment_orders.exists():
        brand.is_active = False
        brand.save()
        return JsonResponse({'success': True, 'message': '브랜드가 비활성화되었습니다. (연결된 주문이 존재)'})

    brand.delete()
    return JsonResponse({'success': True, 'message': '브랜드가 삭제되었습니다.'})


# ============================================================================
# 주문 API
# ============================================================================

@fulfillment_access_required
@require_http_methods(["GET"])
def get_orders(request):
    """주문 목록 JSON API (필터/검색/페이징)"""
    user = request.user
    qs = FulfillmentOrder.objects.select_related('client', 'brand', 'created_by')

    # 고객사 필터 (권한)
    qs = qs.filter(_get_client_filter(user))

    # 필터 파라미터
    client_id = request.GET.get('client_id')
    brand_id = request.GET.get('brand_id')
    platform = request.GET.get('platform')
    status = request.GET.get('status')
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 20)

    if client_id:
        qs = qs.filter(client_id=client_id)
    if brand_id:
        qs = qs.filter(brand_id=brand_id)
    if platform:
        qs = qs.filter(platform=platform)
    if status:
        qs = qs.filter(status=status)
    if search:
        q = (
            Q(product_name__icontains=search) |
            Q(invoice_number__icontains=search)
        )
        # FF-XXXXX 자체코드 검색 지원
        ff_match = re.match(r'^FF-?(\d+)$', search.strip(), re.IGNORECASE)
        if ff_match:
            q |= Q(id=int(ff_match.group(1)))
        qs = qs.filter(q)
    if date_from:
        qs = qs.filter(created_at__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__lte=date_to)

    # 정렬
    qs = qs.order_by('-created_at')

    # 페이징
    try:
        page_size = min(int(page_size), 100)
    except (ValueError, TypeError):
        page_size = 20

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except Exception:
        page_obj = paginator.page(1)

    # platform_data JSON 값 내 검색 (DB 필터 이후 Python 레벨)
    orders_list = list(page_obj)
    if search and not ff_match:
        # DB 쿼리에서 이미 product_name/invoice_number 매칭된 결과가 있으므로,
        # platform_data 검색은 전체 queryset에서 추가 매칭을 위해 별도 처리
        # 간단한 접근: DB 결과에 이미 포함되지 않은 항목을 platform_data에서 추가 검색
        search_lower = search.lower()
        db_ids = {o.id for o in orders_list}

        # platform_data 내에서도 검색하여 추가 매칭
        extra_qs = FulfillmentOrder.objects.select_related('client', 'brand', 'created_by')
        extra_qs = extra_qs.filter(_get_client_filter(user))
        if client_id:
            extra_qs = extra_qs.filter(client_id=client_id)
        if brand_id:
            extra_qs = extra_qs.filter(brand_id=brand_id)
        if platform:
            extra_qs = extra_qs.filter(platform=platform)
        if status:
            extra_qs = extra_qs.filter(status=status)
        if date_from:
            extra_qs = extra_qs.filter(created_at__gte=date_from)
        if date_to:
            extra_qs = extra_qs.filter(created_at__lte=date_to)
        extra_qs = extra_qs.exclude(id__in=db_ids).order_by('-created_at')

        for order in extra_qs[:200]:
            if order.platform_data:
                for val in order.platform_data.values():
                    if search_lower in str(val).lower():
                        orders_list.append(order)
                        break

    orders = []
    for order in orders_list:
        orders.append({
            'id': order.id,
            'internal_code': order.internal_code,
            'client_id': order.client_id,
            'client_name': order.client.company_name,
            'brand_id': order.brand_id,
            'brand_name': order.brand.name if order.brand else '',
            'platform': order.platform,
            'platform_display': order.get_platform_display(),
            'product_name': order.product_name,
            'quantity': order.quantity,
            'box_quantity': order.box_quantity,
            'pallet_quantity': order.pallet_quantity,
            'invoice_number': order.invoice_number,
            'status': order.status,
            'status_display': order.get_status_display(),
            'platform_data': order.platform_data,
            'confirmed_at': timezone.localtime(order.confirmed_at).strftime('%Y-%m-%d %H:%M') if order.confirmed_at else '',
            'shipped_at': timezone.localtime(order.shipped_at).strftime('%Y-%m-%d %H:%M') if order.shipped_at else '',
            'synced_at': timezone.localtime(order.synced_at).strftime('%Y-%m-%d %H:%M') if order.synced_at else '',
            'created_by': order.created_by.name if order.created_by else '',
            'created_at': timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M'),
        })

    return JsonResponse({
        'orders': orders,
        'total': paginator.count,
        'page': page_obj.number,
        'total_pages': paginator.num_pages,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
    })


@fulfillment_access_required
@require_http_methods(["POST"])
def create_order(request):
    """주문 등록"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    user = request.user
    client_id = data.get('client_id')

    # 고객사 검증
    if not client_id:
        return JsonResponse({'error': '거래처를 선택해주세요.'}, status=400)

    try:
        client = Client.objects.get(id=client_id, is_active=True)
    except Client.DoesNotExist:
        return JsonResponse({'error': '유효하지 않은 거래처입니다.'}, status=400)

    # 고객사 사용자는 자기 거래처만 등록 가능
    if user.is_client and not user.clients.filter(id=client_id).exists():
        return JsonResponse({'error': '해당 거래처에 대한 권한이 없습니다.'}, status=403)

    # 브랜드 검증
    brand = None
    brand_id = data.get('brand_id')
    if brand_id:
        try:
            brand = Brand.objects.get(id=brand_id, client=client, is_active=True)
        except Brand.DoesNotExist:
            return JsonResponse({'error': '유효하지 않은 브랜드입니다.'}, status=400)

    # 필수 필드 검증
    platform = data.get('platform', '')
    product_name = data.get('product_name', '').strip()

    if not platform:
        return JsonResponse({'error': '플랫폼을 선택해주세요.'}, status=400)
    if not product_name:
        return JsonResponse({'error': '상품명을 입력해주세요.'}, status=400)

    def safe_int(val, default=0):
        try:
            return int(val or default)
        except (ValueError, TypeError):
            return default

    # 고정 필드와 platform_data 분리
    # 요청 데이터에서 시스템/관계 필드를 제외한 나머지를 분류
    system_keys = {'client_id', 'brand_id', 'platform', 'status'}
    platform_data = data.get('platform_data', {})

    # data에서 고정 필드가 아니고, 시스템 키도 아닌 것들은 platform_data에 추가
    for key, val in data.items():
        if key in system_keys or key in FIXED_MODEL_FIELDS or key == 'platform_data':
            continue
        # 문자열 값이면 platform_data에 넣기
        if isinstance(val, str):
            val = val.strip()
        if val not in (None, ''):
            platform_data[key] = val

    order = FulfillmentOrder.objects.create(
        client=client,
        brand=brand,
        platform=platform,
        product_name=product_name,
        quantity=safe_int(data.get('quantity')),
        box_quantity=safe_int(data.get('box_quantity')),
        pallet_quantity=safe_int(data.get('pallet_quantity')),
        invoice_number=data.get('invoice_number', '').strip(),
        platform_data=platform_data,
        created_by=user,
    )

    # 슬랙 알림
    send_order_created_notification(order)

    return JsonResponse({
        'success': True,
        'message': '주문이 등록되었습니다.',
        'order_id': order.id,
    })


@fulfillment_access_required
@require_http_methods(["POST"])
def bulk_paste_orders(request):
    """
    붙여넣기 벌크 주문 등록

    PlatformColumnConfig 기반으로 컬럼 순서를 동적으로 결정합니다.
    컬럼 순서: 고정 필드(product_name, quantity, box_quantity, pallet_quantity, invoice_number)
              + 커스텀 필드(PlatformColumnConfig의 display_order 순)
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    user = request.user
    client_id = data.get('client_id')
    brand_id = data.get('brand_id')
    platform = data.get('platform', '')
    paste_text = data.get('paste_text', '')

    if not client_id:
        return JsonResponse({'error': '거래처를 선택해주세요.'}, status=400)
    if not platform:
        return JsonResponse({'error': '플랫폼을 선택해주세요.'}, status=400)
    if not paste_text.strip():
        return JsonResponse({'error': '붙여넣기 데이터가 없습니다.'}, status=400)

    try:
        client = Client.objects.get(id=client_id, is_active=True)
    except Client.DoesNotExist:
        return JsonResponse({'error': '유효하지 않은 거래처입니다.'}, status=400)

    # 고객사 권한 체크
    if user.is_client and not user.clients.filter(id=client_id).exists():
        return JsonResponse({'error': '해당 거래처에 대한 권한이 없습니다.'}, status=403)

    # 브랜드
    brand = None
    if brand_id:
        try:
            brand = Brand.objects.get(id=brand_id, client=client, is_active=True)
        except Brand.DoesNotExist:
            pass

    # PlatformColumnConfig에서 현재 플랫폼의 활성 컬럼 로드 (display_order 순)
    custom_columns = list(PlatformColumnConfig.objects.filter(
        platform=platform, is_active=True
    ).order_by('display_order'))

    # 컬럼 순서 결정: 고정 필드 + 커스텀 필드
    # 고정 필드 정의 (key, label, type)
    fixed_column_defs = [
        ('product_name', '상품명', 'text'),
        ('quantity', '수량', 'number'),
        ('box_quantity', '박스수', 'number'),
        ('pallet_quantity', '팔레트수', 'number'),
        ('invoice_number', '송장번호', 'text'),
    ]

    # 전체 컬럼 맵핑 순서: 고정 필드 + 커스텀 필드
    column_order = []
    for key, label, col_type in fixed_column_defs:
        column_order.append({'key': key, 'type': col_type, 'is_fixed': True})
    for cc in custom_columns:
        column_order.append({'key': cc.key, 'type': cc.column_type, 'is_fixed': False})

    # 탭 구분 텍스트 파싱
    lines = paste_text.strip().split('\n')
    created_count = 0
    errors = []

    def safe_int(val, default=0):
        """숫자 파싱 - 쉼표 제거 후 정수 변환"""
        if not val:
            return default
        try:
            cleaned = str(val).strip().replace(',', '')
            return int(cleaned) if cleaned else default
        except (ValueError, TypeError):
            return default

    for row_idx, line in enumerate(lines, 1):
        line = line.strip()
        if not line:
            continue

        cols = line.split('\t')

        try:
            model_fields = {}
            platform_data = {}

            for col_idx, col_def in enumerate(column_order):
                value = cols[col_idx].strip() if col_idx < len(cols) else ''
                key = col_def['key']

                if col_def['is_fixed']:
                    # 고정 필드: 모델에 직접 설정
                    if col_def['type'] == 'number':
                        model_fields[key] = safe_int(value)
                    else:
                        model_fields[key] = value
                else:
                    # 커스텀 필드: platform_data에 저장
                    if value:
                        platform_data[key] = value

            product_name = model_fields.get('product_name', '')
            if not product_name:
                errors.append(f'{row_idx}행: 상품명이 비어있습니다.')
                continue

            FulfillmentOrder.objects.create(
                client=client,
                brand=brand,
                platform=platform,
                product_name=product_name,
                quantity=model_fields.get('quantity', 0),
                box_quantity=model_fields.get('box_quantity', 0),
                pallet_quantity=model_fields.get('pallet_quantity', 0),
                invoice_number=model_fields.get('invoice_number', ''),
                platform_data=platform_data,
                created_by=user,
            )
            created_count += 1

        except Exception as e:
            errors.append(f'{row_idx}행: {str(e)}')

    result = {
        'success': True,
        'message': f'{created_count}건이 등록되었습니다.',
        'created_count': created_count,
    }
    if errors:
        result['errors'] = errors[:20]
        result['error_count'] = len(errors)

    # 슬랙 알림 (1건 이상 등록 시 요약 메시지 1건만 발송)
    if created_count > 0:
        send_bulk_orders_notification(
            client=client,
            brand=brand,
            platform=platform,
            created_count=created_count,
            error_count=len(errors),
            user=user,
        )

    return JsonResponse(result)


@admin_required
@require_http_methods(["POST"])
def update_order(request, order_id):
    """주문 수정 (관리자 전용)"""
    try:
        order = FulfillmentOrder.objects.get(id=order_id)
    except FulfillmentOrder.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    # 거래처 변경
    if 'client_id' in data:
        try:
            client = Client.objects.get(id=data['client_id'], is_active=True)
            order.client = client
        except Client.DoesNotExist:
            return JsonResponse({'error': '유효하지 않은 거래처입니다.'}, status=400)

    # 브랜드 변경
    if 'brand_id' in data:
        if data['brand_id']:
            try:
                brand = Brand.objects.get(id=data['brand_id'], client=order.client, is_active=True)
                order.brand = brand
            except Brand.DoesNotExist:
                pass
        else:
            order.brand = None

    # 플랫폼 변경
    if 'platform' in data:
        val = data['platform']
        order.platform = val.strip() if isinstance(val, str) else val

    # 고정 텍스트 필드
    fixed_text_fields = ['product_name', 'invoice_number']
    for field in fixed_text_fields:
        if field in data:
            val = data[field]
            setattr(order, field, val.strip() if isinstance(val, str) else val)

    # 고정 숫자 필드
    fixed_int_fields = ['quantity', 'box_quantity', 'pallet_quantity']
    for field in fixed_int_fields:
        if field in data:
            try:
                setattr(order, field, int(data[field] or 0))
            except (ValueError, TypeError):
                pass

    # platform_data 업데이트
    # platform_data를 직접 전달하면 그대로 사용
    if 'platform_data' in data:
        order.platform_data = data['platform_data']
    else:
        # data에서 고정/시스템 필드가 아닌 것들은 platform_data에 병합
        system_keys = {'client_id', 'brand_id', 'platform', 'status', 'platform_data'}
        current_pd = order.platform_data or {}
        for key, val in data.items():
            if key in system_keys or key in FIXED_MODEL_FIELDS:
                continue
            if isinstance(val, str):
                val = val.strip()
            current_pd[key] = val
        order.platform_data = current_pd

    order.save()

    return JsonResponse({
        'success': True,
        'message': '주문이 수정되었습니다.',
    })


@admin_required
@require_http_methods(["POST"])
def delete_order(request, order_id):
    """주문 삭제 (관리자 전용)"""
    try:
        order = FulfillmentOrder.objects.get(id=order_id)
    except FulfillmentOrder.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)

    order.delete()
    return JsonResponse({
        'success': True,
        'message': '주문이 삭제되었습니다.',
    })


@admin_or_worker_required
@require_http_methods(["POST"])
def update_status(request, order_id):
    """상태 변경 (관리자/작업자 전용) - 확인/출고/전산반영"""
    try:
        order = FulfillmentOrder.objects.select_related(
            'created_by', 'client', 'shipped_by',
        ).get(id=order_id)
    except FulfillmentOrder.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    action = data.get('action', '')
    user = request.user

    status_actions = {
        'confirm': {
            'method': order.confirm,
            'message': '확인완료 처리되었습니다.',
            'system_msg': '상태가 [확인완료]로 변경되었습니다.',
            'time_field': 'confirmed_at',
        },
        'ship': {
            'method': order.ship,
            'message': '출고완료 처리되었습니다.',
            'system_msg': '상태가 [출고완료]로 변경되었습니다.',
            'time_field': 'shipped_at',
        },
        'sync': {
            'method': order.sync,
            'message': '전산반영 처리되었습니다.',
            'system_msg': '상태가 [전산반영]으로 변경되었습니다.',
            'time_field': 'synced_at',
        },
    }

    if action not in status_actions:
        return JsonResponse({'error': '잘못된 액션입니다.'}, status=400)

    cfg = status_actions[action]

    # 출고처리 시 박스수/팔레트수/송장번호 업데이트
    if action == 'ship':
        ship_data = data.get('ship_data', {})
        update_fields = []
        if 'box_quantity' in ship_data:
            try:
                order.box_quantity = int(ship_data['box_quantity']) if ship_data['box_quantity'] not in (None, '') else 0
                update_fields.append('box_quantity')
            except (ValueError, TypeError):
                pass
        if 'pallet_quantity' in ship_data:
            try:
                order.pallet_quantity = int(ship_data['pallet_quantity']) if ship_data['pallet_quantity'] not in (None, '') else 0
                update_fields.append('pallet_quantity')
            except (ValueError, TypeError):
                pass
        if 'invoice_number' in ship_data:
            order.invoice_number = str(ship_data['invoice_number'] or '')
            update_fields.append('invoice_number')
        if update_fields:
            order.save(update_fields=update_fields + ['updated_at'])

    if cfg['method'](user):
        # 시스템 댓글 자동 추가
        FulfillmentComment.objects.create(
            order=order,
            author=user,
            content=f"{cfg['system_msg']} ({user.name})",
            is_system=True,
        )

        # 출고완료 시 등록자에게 이메일 알림 (백그라운드)
        if action == 'ship':
            order.shipped_by = user
            send_shipment_notification_async(order)

        time_val = getattr(order, cfg['time_field'])
        return JsonResponse({
            'success': True,
            'message': cfg['message'],
            'status': order.status,
            'status_display': order.get_status_display(),
            cfg['time_field']: timezone.localtime(time_val).strftime('%Y-%m-%d %H:%M') if time_val else '',
        })
    else:
        error_msgs = {
            'confirm': '확인 처리할 수 없는 상태입니다.',
            'ship': '출고 처리할 수 없는 상태입니다.',
            'sync': '전산반영 처리할 수 없는 상태입니다.',
        }
        return JsonResponse({'error': error_msgs[action]}, status=400)


@admin_or_worker_required
@require_http_methods(["POST"])
def bulk_update_status(request):
    """일괄 상태 변경 (관리자/작업자 전용) - 확인/출고/전산반영"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    order_ids = data.get('order_ids', [])
    action = data.get('action', '')
    # 출고처리 시 주문별 박스수/팔레트수/송장번호 데이터 {order_id: {box_quantity, pallet_quantity, invoice_number}}
    ship_data_map = data.get('ship_data_map', {})

    if not order_ids:
        return JsonResponse({'error': '선택된 주문이 없습니다.'}, status=400)

    action_map = {
        'confirm': {
            'method': 'confirm',
            'label': '확인완료',
            'system_msg': '상태가 [확인완료]로 변경되었습니다.',
        },
        'ship': {
            'method': 'ship',
            'label': '출고완료',
            'system_msg': '상태가 [출고완료]로 변경되었습니다.',
        },
        'sync': {
            'method': 'sync',
            'label': '전산반영',
            'system_msg': '상태가 [전산반영]으로 변경되었습니다.',
        },
    }

    if action not in action_map:
        return JsonResponse({'error': '잘못된 액션입니다.'}, status=400)

    cfg = action_map[action]
    user = request.user
    orders = FulfillmentOrder.objects.select_related(
        'created_by', 'client', 'shipped_by',
    ).filter(id__in=order_ids)

    success_count = 0
    fail_count = 0
    shipped_orders = []

    for order in orders:
        # 출고처리 시 박스수/팔레트수/송장번호 업데이트
        if action == 'ship':
            ship_data = ship_data_map.get(str(order.id), {})
            update_fields = []
            if 'box_quantity' in ship_data:
                try:
                    order.box_quantity = int(ship_data['box_quantity']) if ship_data['box_quantity'] not in (None, '') else 0
                    update_fields.append('box_quantity')
                except (ValueError, TypeError):
                    pass
            if 'pallet_quantity' in ship_data:
                try:
                    order.pallet_quantity = int(ship_data['pallet_quantity']) if ship_data['pallet_quantity'] not in (None, '') else 0
                    update_fields.append('pallet_quantity')
                except (ValueError, TypeError):
                    pass
            if 'invoice_number' in ship_data:
                order.invoice_number = str(ship_data['invoice_number'] or '')
                update_fields.append('invoice_number')
            if update_fields:
                order.save(update_fields=update_fields + ['updated_at'])

        method = getattr(order, cfg['method'])
        if method(user):
            FulfillmentComment.objects.create(
                order=order,
                author=user,
                content=f"{cfg['system_msg']} ({user.name}) [일괄처리]",
                is_system=True,
            )

            if action == 'ship':
                order.shipped_by = user
                shipped_orders.append(order)

            success_count += 1
        else:
            fail_count += 1

    # 출고완료 건 일괄 이메일 알림 (백그라운드)
    if shipped_orders:
        send_shipment_notifications_async(shipped_orders)

    msg = f'{success_count}건 {cfg["label"]} 처리되었습니다.'
    if fail_count:
        msg += f' ({fail_count}건은 상태 조건 불일치로 제외)'

    return JsonResponse({
        'success': True,
        'message': msg,
        'success_count': success_count,
        'fail_count': fail_count,
    })


@fulfillment_access_required
@require_http_methods(["GET"])
def export_excel(request):
    """엑셀 다운로드 (현재 필터 조건 적용)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return JsonResponse({'error': 'openpyxl 패키지가 필요합니다.'}, status=500)

    user = request.user
    qs = FulfillmentOrder.objects.select_related('client', 'brand')

    # 고객사 필터 (권한)
    qs = qs.filter(_get_client_filter(user))

    # 필터 파라미터 (목록 API와 동일)
    client_id = request.GET.get('client_id')
    brand_id = request.GET.get('brand_id')
    platform = request.GET.get('platform')
    status = request.GET.get('status')
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if client_id:
        qs = qs.filter(client_id=client_id)
    if brand_id:
        qs = qs.filter(brand_id=brand_id)
    if platform:
        qs = qs.filter(platform=platform)
    if status:
        qs = qs.filter(status=status)
    if search:
        q = (
            Q(product_name__icontains=search) |
            Q(invoice_number__icontains=search)
        )
        ff_match = re.match(r'^FF-?(\d+)$', search.strip(), re.IGNORECASE)
        if ff_match:
            q |= Q(id=int(ff_match.group(1)))
        qs = qs.filter(q)
    if date_from:
        qs = qs.filter(created_at__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__lte=date_to)

    qs = qs.order_by('-created_at')

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출고 주문 목록'

    # 헤더 스타일
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 플랫폼별 커스텀 컬럼 조회
    platform_cols = []
    if platform:
        platform_cols = list(PlatformColumnConfig.objects.filter(
            platform=platform, is_active=True
        ).order_by('display_order'))

    # 헤더: 고정 필드 + 상태 + 커스텀 컬럼
    headers = [
        '자체코드', '거래처', '브랜드', '플랫폼',
        '상품명', '수량', '박스수', '팔레트수', '송장번호',
        '상태',
    ]
    # 플랫폼 커스텀 컬럼 헤더 추가
    for pc in platform_cols:
        headers.append(pc.name)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터
    for row_idx, order in enumerate(qs, 2):
        row_data = [
            order.internal_code,
            order.client.company_name,
            order.brand.name if order.brand else '',
            order.get_platform_display(),
            order.product_name,
            order.quantity,
            order.box_quantity,
            order.pallet_quantity,
            order.invoice_number,
            order.get_status_display(),
        ]
        # 플랫폼 커스텀 컬럼 데이터 추가
        for pc in platform_cols:
            row_data.append(
                order.platform_data.get(pc.key, '') if order.platform_data else ''
            )

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 컬럼 너비 조정
    col_widths = [12, 15, 12, 12, 30, 10, 10, 10, 15, 10]
    # 플랫폼 커스텀 컬럼 너비 추가
    for _ in platform_cols:
        col_widths.append(15)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 응답
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="fulfillment_orders_{now_str}.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 댓글 API
# ============================================================================

def _check_order_access(user, order):
    """주문에 대한 접근 권한 확인"""
    if user.is_admin or user.is_superuser or user.is_worker:
        return True
    if user.is_client:
        return user.clients.filter(id=order.client_id).exists()
    return False


@fulfillment_access_required
@require_http_methods(["GET"])
def get_comments(request, order_id):
    """댓글 목록 조회"""
    try:
        order = FulfillmentOrder.objects.get(id=order_id)
    except FulfillmentOrder.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)

    # 권한 확인
    if not _check_order_access(request.user, order):
        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)

    comments = order.comments.select_related('author').all()
    comment_list = []
    for c in comments:
        comment_data = {
            'id': c.id,
            'author_name': c.author.name if c.author else '시스템',
            'author_role': c.author.get_role_display() if c.author else '',
            'author_id': c.author.id if c.author else None,
            'content': c.content,
            'is_system': c.is_system,
            'created_at': timezone.localtime(c.created_at).strftime('%Y-%m-%d %H:%M'),
            'is_mine': c.author_id == request.user.id if c.author else False,
        }
        # 첨부파일 정보
        if c.file:
            ext = os.path.splitext(c.file.name)[1].lower()
            comment_data['file_url'] = c.file.url
            comment_data['file_name'] = os.path.basename(c.file.name)
            try:
                comment_data['file_size'] = c.file.size
            except Exception:
                comment_data['file_size'] = 0
            comment_data['is_image'] = ext in COMMENT_IMAGE_EXTENSIONS
        comment_list.append(comment_data)

    return JsonResponse({'comments': comment_list})


@fulfillment_access_required
@require_http_methods(["POST"])
def add_comment(request, order_id):
    """댓글 등록 (텍스트 + 선택적 파일 첨부)"""
    try:
        order = FulfillmentOrder.objects.get(id=order_id)
    except FulfillmentOrder.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)

    # 권한 확인
    if not _check_order_access(request.user, order):
        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)

    # multipart/form-data 또는 JSON 모두 지원
    content_type = request.content_type or ''
    if 'multipart/form-data' in content_type:
        content = request.POST.get('content', '').strip()
        uploaded_file = request.FILES.get('file')
    else:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)
        content = data.get('content', '').strip()
        uploaded_file = None

    if not content and not uploaded_file:
        return JsonResponse({'error': '댓글 내용을 입력하거나 파일을 첨부해주세요.'}, status=400)

    # 파일 검증
    if uploaded_file:
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        if ext not in COMMENT_ALLOWED_EXTENSIONS:
            allowed = ', '.join(sorted(COMMENT_ALLOWED_EXTENSIONS))
            return JsonResponse(
                {'error': f'허용되지 않는 파일 형식입니다. ({allowed})'},
                status=400,
            )
        if uploaded_file.size > COMMENT_MAX_FILE_SIZE:
            return JsonResponse({'error': '파일 크기는 10MB 이하여야 합니다.'}, status=400)

    comment = FulfillmentComment.objects.create(
        order=order,
        author=request.user,
        content=content or '',
    )

    if uploaded_file:
        comment.file = uploaded_file
        comment.save(update_fields=['file'])

    # 응답 데이터
    response_data = {
        'id': comment.id,
        'author_name': comment.author.name,
        'author_role': comment.author.get_role_display(),
        'author_id': comment.author.id,
        'content': comment.content,
        'is_system': False,
        'created_at': timezone.localtime(comment.created_at).strftime('%Y-%m-%d %H:%M'),
        'is_mine': True,
    }
    if comment.file:
        file_ext = os.path.splitext(comment.file.name)[1].lower()
        response_data['file_url'] = comment.file.url
        response_data['file_name'] = os.path.basename(comment.file.name)
        response_data['file_size'] = comment.file.size
        response_data['is_image'] = file_ext in COMMENT_IMAGE_EXTENSIONS

    return JsonResponse({
        'success': True,
        'message': '댓글이 등록되었습니다.',
        'comment': response_data,
    })


@fulfillment_access_required
@require_http_methods(["POST"])
def delete_comment(request, order_id, comment_id):
    """댓글 삭제 (본인 또는 관리자)"""
    try:
        comment = FulfillmentComment.objects.get(id=comment_id, order_id=order_id)
    except FulfillmentComment.DoesNotExist:
        return JsonResponse({'error': '댓글을 찾을 수 없습니다.'}, status=404)

    user = request.user
    # 본인 댓글이거나 관리자만 삭제 가능
    if comment.author_id != user.id and not (user.is_admin or user.is_superuser):
        return JsonResponse({'error': '삭제 권한이 없습니다.'}, status=403)

    # 첨부파일 삭제
    if comment.file:
        try:
            comment.file.delete(save=False)
        except Exception:
            pass  # 파일 삭제 실패해도 댓글은 삭제

    comment.delete()
    return JsonResponse({
        'success': True,
        'message': '댓글이 삭제되었습니다.',
    })


# ============================================================================
# 플랫폼 컬럼 설정 API
# ============================================================================

@fulfillment_access_required
@require_http_methods(["GET"])
def get_platform_columns(request):
    """플랫폼별 커스텀 컬럼 목록 조회"""
    platform = request.GET.get('platform', '')
    if not platform:
        return JsonResponse({'columns': []})

    active_only = request.GET.get('active_only', 'true') == 'true'

    qs = PlatformColumnConfig.objects.filter(platform=platform)
    if active_only:
        qs = qs.filter(is_active=True)

    columns = []
    for col in qs:
        columns.append({
            'id': col.id,
            'platform': col.platform,
            'name': col.name,
            'key': col.key,
            'column_type': col.column_type,
            'display_order': col.display_order,
            'is_required': col.is_required,
            'is_active': col.is_active,
        })

    return JsonResponse({'columns': columns})


@admin_required
@require_http_methods(["POST"])
def save_platform_columns(request):
    """플랫폼 컬럼 설정 일괄 저장 (관리자 전용)"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    platform = data.get('platform', '')
    columns = data.get('columns', [])
    delete_ids = data.get('delete_ids', [])

    if not platform:
        return JsonResponse({'error': '플랫폼을 선택해주세요.'}, status=400)

    # 삭제
    if delete_ids:
        PlatformColumnConfig.objects.filter(
            id__in=delete_ids, platform=platform
        ).delete()

    # 생성/수정
    for idx, col_data in enumerate(columns):
        col_id = col_data.get('id')
        name = col_data.get('name', '').strip()
        key = col_data.get('key', '').strip()

        if not name or not key:
            continue

        defaults = {
            'name': name,
            'column_type': col_data.get('column_type', 'text'),
            'display_order': col_data.get('display_order', idx),
            'is_required': col_data.get('is_required', False),
            'is_active': col_data.get('is_active', True),
        }

        if col_id:
            PlatformColumnConfig.objects.filter(
                id=col_id, platform=platform
            ).update(**defaults, key=key)
        else:
            if not PlatformColumnConfig.objects.filter(
                platform=platform, key=key
            ).exists():
                PlatformColumnConfig.objects.create(
                    platform=platform, key=key, **defaults
                )

    return JsonResponse({
        'success': True,
        'message': '컬럼 설정이 저장되었습니다.',
    })


# ============================================================================
# 주문 등록 페이지
# ============================================================================

@fulfillment_access_required
def order_create_page(request):
    """주문 등록 페이지 (독립 페이지)"""
    user = request.user

    if user.is_client:
        clients = user.clients.filter(is_active=True)
    else:
        clients = Client.objects.filter(is_active=True).order_by('company_name')

    context = {
        'clients': clients,
        'platforms': FulfillmentOrder.Platform.choices,
        'is_admin': user.is_admin or user.is_superuser,
        'is_client': user.is_client,
        'is_worker': user.is_worker,
    }
    return render(request, 'fulfillment/order_create.html', context)


@fulfillment_access_required
@require_http_methods(["POST"])
def bulk_create_orders(request):
    """다건 주문 등록 (행 단위 테이블 입력)"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    user = request.user
    client_id = data.get('client_id')
    brand_id = data.get('brand_id')
    platform = data.get('platform', '')
    orders_data = data.get('orders', [])

    if not client_id:
        return JsonResponse({'error': '거래처를 선택해주세요.'}, status=400)
    if not platform:
        return JsonResponse({'error': '플랫폼을 선택해주세요.'}, status=400)
    if not orders_data:
        return JsonResponse({'error': '등록할 주문이 없습니다.'}, status=400)

    try:
        client = Client.objects.get(id=client_id, is_active=True)
    except Client.DoesNotExist:
        return JsonResponse({'error': '유효하지 않은 거래처입니다.'}, status=400)

    if user.is_client and not user.clients.filter(id=client_id).exists():
        return JsonResponse({'error': '해당 거래처에 대한 권한이 없습니다.'}, status=403)

    brand = None
    if brand_id:
        try:
            brand = Brand.objects.get(id=brand_id, client=client, is_active=True)
        except Brand.DoesNotExist:
            pass

    def safe_int(val, default=0):
        try:
            return int(str(val).strip().replace(',', '') or default)
        except (ValueError, TypeError):
            return default

    created_count = 0
    errors = []

    for idx, row in enumerate(orders_data, 1):
        product_name = (row.get('product_name') or '').strip()

        if not product_name:
            errors.append(f'{idx}행: 상품명이 비어있습니다.')
            continue

        try:
            # 고정 필드 추출
            model_kwargs = {
                'client': client,
                'brand': brand,
                'platform': platform,
                'product_name': product_name,
                'quantity': safe_int(row.get('quantity')),
                'box_quantity': safe_int(row.get('box_quantity')),
                'pallet_quantity': safe_int(row.get('pallet_quantity')),
                'invoice_number': (row.get('invoice_number') or '').strip(),
                'created_by': user,
            }

            # 나머지 필드는 platform_data에 저장
            platform_data = row.get('platform_data') or {}
            system_keys = {'client_id', 'brand_id', 'platform', 'status', 'platform_data'}
            for key, val in row.items():
                if key in system_keys or key in FIXED_MODEL_FIELDS:
                    continue
                if isinstance(val, str):
                    val = val.strip()
                if val not in (None, ''):
                    platform_data[key] = val

            model_kwargs['platform_data'] = platform_data

            FulfillmentOrder.objects.create(**model_kwargs)
            created_count += 1
        except Exception as e:
            errors.append(f'{idx}행: {str(e)}')

    result = {
        'success': True,
        'message': f'{created_count}건이 등록되었습니다.',
        'created_count': created_count,
    }
    if errors:
        result['errors'] = errors[:20]
        result['error_count'] = len(errors)

    if created_count > 0:
        send_bulk_orders_notification(
            client=client,
            brand=brand,
            platform=platform,
            created_count=created_count,
            error_count=len(errors),
            user=user,
        )

    return JsonResponse(result)
