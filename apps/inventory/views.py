"""
재고 관리 뷰
"""
import json
import logging
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from .models import Product, Location, InventorySession, InventoryRecord

logger = logging.getLogger(__name__)


# ============================================================================
# 권한 체크 데코레이터
# ============================================================================

def admin_required(view_func):
    """관리자 전용 뷰 데코레이터"""
    def wrapper(request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_superuser):
            return JsonResponse({'error': '관리자 권한이 필요합니다.'}, status=403)
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    wrapper.__doc__ = view_func.__doc__
    return wrapper


def staff_required(view_func):
    """관리자 또는 작업자 접근 가능 데코레이터"""
    def wrapper(request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_worker or request.user.is_superuser):
            return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    wrapper.__doc__ = view_func.__doc__
    return wrapper


# ============================================================================
# 페이지 뷰
# ============================================================================

@login_required
@staff_required
def session_page(request):
    """재고조사 세션 관리 페이지"""
    active_session = InventorySession.objects.filter(status='active').first()
    return render(request, 'inventory/session.html', {
        'active_session': active_session,
    })


@login_required
@staff_required
def scan_page(request):
    """재고 스캔 입력 페이지 (관리자 + 작업자)"""
    active_session = InventorySession.objects.filter(status='active').first()
    return render(request, 'inventory/scan.html', {
        'active_session': active_session,
    })


@login_required
@staff_required
def status_page(request):
    """재고 현황 조회 페이지 (관리자 + 작업자)"""
    sessions = InventorySession.objects.all()
    active_session = sessions.filter(status='active').first()
    return render(request, 'inventory/status.html', {
        'sessions': sessions,
        'active_session': active_session,
    })


@login_required
@staff_required
def products_page(request):
    """상품 마스터 관리 페이지"""
    return render(request, 'inventory/products.html')


# ============================================================================
# API: 상품 마스터
# ============================================================================

def _get_chosung_regex(chosung_char):
    """초성 문자에 해당하는 한글 정규식 문자 클래스를 반환한다.

    예: 'ㄱ' → '^[가-깋]', 'ㄷ' → '^[다-딯]', 'ㅎ' → '^[하-힣]'
    PostgreSQL에서 locale-independent하게 동작하기 위해 정규식 사용.
    """
    CHO_LIST = 'ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ'
    idx = CHO_LIST.find(chosung_char)
    if idx < 0:
        return None

    # 한글 유니코드: 0xAC00 + (초성idx * 21 * 28)
    start = chr(0xAC00 + idx * 21 * 28)
    end = chr(0xAC00 + idx * 21 * 28 + 21 * 28 - 1)
    return f'^[{start}-{end}]'


# 초성 버튼 → 실제 검색할 초성 목록 매핑
# 예: 'ㄱ' → ['ㄱ', 'ㄲ'], 'ㄷ' → ['ㄷ', 'ㄸ'], 'ㅂ' → ['ㅂ', 'ㅃ'], 'ㅅ' → ['ㅅ', 'ㅆ'], 'ㅈ' → ['ㅈ', 'ㅉ']
CHOSUNG_GROUP = {
    'ㄱ': ['ㄱ', 'ㄲ'], 'ㄴ': ['ㄴ'], 'ㄷ': ['ㄷ', 'ㄸ'], 'ㄹ': ['ㄹ'],
    'ㅁ': ['ㅁ'], 'ㅂ': ['ㅂ', 'ㅃ'], 'ㅅ': ['ㅅ', 'ㅆ'], 'ㅇ': ['ㅇ'],
    'ㅈ': ['ㅈ', 'ㅉ'], 'ㅊ': ['ㅊ'], 'ㅋ': ['ㅋ'], 'ㅌ': ['ㅌ'],
    'ㅍ': ['ㅍ'], 'ㅎ': ['ㅎ'],
}


@login_required
@staff_required
@require_GET
def get_products(request):
    """상품 목록을 조회한다.

    Query params:
        search: 검색어 (바코드 or 상품명)
        initial: 초성 또는 알파벳 (ㄱ~ㅎ, A~Z) — 해당 글자로 시작하는 상품 검색
    """
    from django.db.models import Q

    search = request.GET.get('search', '').strip()
    initial = request.GET.get('initial', '').strip()
    products = Product.objects.all()

    if search:
        products = products.filter(
            Q(barcode__icontains=search) |
            Q(name__icontains=search) |
            Q(display_name__icontains=search)
        )
    elif initial:
        # 초성 검색 (ㄱ~ㅎ)
        if initial in CHOSUNG_GROUP:
            q = Q()
            for cho in CHOSUNG_GROUP[initial]:
                regex = _get_chosung_regex(cho)
                if regex:
                    q |= Q(name__regex=regex) | Q(display_name__regex=regex)
            products = products.filter(q)
        # 알파벳 검색 (A~Z)
        elif initial.isalpha() and len(initial) == 1:
            products = products.filter(
                Q(name__iregex=f'^{initial}') |
                Q(display_name__iregex=f'^{initial}')
            )
        # 숫자 검색 (0~9)
        elif initial == '0':
            products = products.filter(
                Q(name__regex=r'^[0-9]') |
                Q(display_name__regex=r'^[0-9]')
            )

    products = products.order_by('name')[:100]

    return JsonResponse({
        'products': [
            {
                'id': p.pk,
                'barcode': p.barcode,
                'name': p.name,
                'display_name': p.display_name,
                'created_at': timezone.localtime(p.created_at).strftime('%Y-%m-%d %H:%M'),
            }
            for p in products
        ],
        'total': Product.objects.count(),
    })


@csrf_exempt
@login_required
@staff_required
@require_POST
def create_product(request):
    """상품을 수동으로 등록한다."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    barcode = data.get('barcode', '').strip()
    name = data.get('name', '').strip()
    display_name = data.get('display_name', '').strip()

    if not barcode or not name:
        return JsonResponse({'error': '바코드와 상품명을 모두 입력해주세요.'}, status=400)

    if Product.objects.filter(barcode=barcode, name=name).exists():
        return JsonResponse({'error': f'이미 등록된 상품입니다: {barcode} / {name}'}, status=400)

    product = Product.objects.create(barcode=barcode, name=name, display_name=display_name)

    return JsonResponse({
        'success': True,
        'product': {
            'id': product.pk,
            'barcode': product.barcode,
            'name': product.name,
            'display_name': product.display_name,
        },
    })


@csrf_exempt
@login_required
@staff_required
@require_POST
def update_product(request, product_id):
    """상품 정보를 수정한다."""
    try:
        product = Product.objects.get(pk=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'error': '상품을 찾을 수 없습니다.'}, status=404)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    barcode = data.get('barcode', '').strip()
    name = data.get('name', '').strip()
    display_name = data.get('display_name', '').strip()

    if not barcode or not name:
        return JsonResponse({'error': '바코드와 상품명을 모두 입력해주세요.'}, status=400)

    # 바코드+상품명 중복 체크 (자기 자신 제외)
    if Product.objects.filter(barcode=barcode, name=name).exclude(pk=product_id).exists():
        return JsonResponse({'error': f'이미 등록된 상품입니다: {barcode} / {name}'}, status=400)

    product.barcode = barcode
    product.name = name
    product.display_name = display_name
    product.save(update_fields=['barcode', 'name', 'display_name', 'updated_at'])

    return JsonResponse({
        'success': True,
        'product': {
            'id': product.pk,
            'barcode': product.barcode,
            'name': product.name,
            'display_name': product.display_name,
        },
    })


@csrf_exempt
@login_required
@staff_required
def delete_product(request, product_id):
    """상품을 삭제한다."""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE 메서드만 허용됩니다.'}, status=405)

    try:
        product = Product.objects.get(pk=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'error': '상품을 찾을 수 없습니다.'}, status=404)

    product.delete()
    return JsonResponse({'success': True})


@csrf_exempt
@login_required
@staff_required
@require_POST
def upload_products_excel(request):
    """엑셀 파일로 상품을 일괄 등록한다.

    엑셀 첫 번째 행: 헤더 (바코드, 상품명)
    두 번째 행부터: 데이터
    """
    import openpyxl
    import xlrd

    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'error': '파일을 선택해주세요.'}, status=400)

    filename = excel_file.name.lower()
    if not filename.endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'}, status=400)

    try:
        # 엑셀 파싱
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        else:
            content = excel_file.read()
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]
            rows = []
            for row_idx in range(1, ws.nrows):
                rows.append(tuple(ws.cell_value(row_idx, col) for col in range(ws.ncols)))

        # 헤더에서 바코드/상품명/관리명 컬럼 인덱스 찾기
        barcode_idx = _find_column_index(headers, ['바코드', '바코드번호', 'barcode', 'BarCode', 'BARCODE'])
        name_idx = _find_column_index(headers, ['상품명', '품명', '제품명', '이름', 'name', 'product_name', '상품이름'])
        display_name_idx = _find_column_index(headers, ['관리명', '관리이름', '관리상품명', 'display_name', '별칭'])

        if barcode_idx is None:
            return JsonResponse({'error': '바코드 컬럼을 찾을 수 없습니다. 헤더에 "바코드" 또는 "barcode"가 포함되어야 합니다.'}, status=400)
        if name_idx is None:
            return JsonResponse({'error': '상품명 컬럼을 찾을 수 없습니다. 헤더에 "상품명" 또는 "name"이 포함되어야 합니다.'}, status=400)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for row_num, row in enumerate(rows, start=2):
            barcode_val = str(row[barcode_idx] or '').strip() if barcode_idx < len(row) else ''
            name_val = str(row[name_idx] or '').strip() if name_idx < len(row) else ''
            display_name_val = ''
            if display_name_idx is not None and display_name_idx < len(row):
                display_name_val = str(row[display_name_idx] or '').strip()

            # 바코드가 숫자로 읽힌 경우 정수로 변환
            if barcode_val and '.' in barcode_val:
                try:
                    barcode_val = str(int(float(barcode_val)))
                except (ValueError, OverflowError):
                    pass

            if not barcode_val or not name_val:
                skipped_count += 1
                continue

            try:
                if display_name_val:
                    product, created = Product.objects.update_or_create(
                        barcode=barcode_val,
                        name=name_val,
                        defaults={'display_name': display_name_val},
                    )
                else:
                    product, created = Product.objects.get_or_create(
                        barcode=barcode_val,
                        name=name_val,
                    )
                if created:
                    created_count += 1
                else:
                    updated_count += 1
            except Exception as e:
                errors.append(f'{row_num}행: {str(e)}')
                skipped_count += 1

        result_msg = f'신규 {created_count}건, 업데이트 {updated_count}건'
        if skipped_count:
            result_msg += f', 스킵 {skipped_count}건'

        return JsonResponse({
            'success': True,
            'message': result_msg,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': errors[:10],  # 최대 10개만
        })

    except Exception as e:
        logger.error('상품 엑셀 업로드 실패: %s', e, exc_info=True)
        return JsonResponse({'error': f'파일 처리 중 오류: {str(e)}'}, status=500)


@login_required
@staff_required
@require_GET
def lookup_product(request):
    """바코드로 상품을 조회한다 (스캔 시 자동완성용).

    동일 바코드에 여러 상품이 등록되어 있으면 복수 결과를 반환한다.

    Query params:
        barcode: 상품 바코드
    """
    barcode = request.GET.get('barcode', '').strip()
    if not barcode:
        return JsonResponse({'found': False})

    products = Product.objects.filter(barcode=barcode).order_by('name')

    if not products.exists():
        return JsonResponse({'found': False})

    product_list = [
        {
            'id': p.pk,
            'barcode': p.barcode,
            'name': p.name,
            'display_name': p.display_name,
        }
        for p in products
    ]

    if len(product_list) == 1:
        # 단일 상품 — 기존 호환
        return JsonResponse({
            'found': True,
            'multiple': False,
            'product': product_list[0],
        })
    else:
        # 동일 바코드 복수 상품 — 선택 필요
        return JsonResponse({
            'found': True,
            'multiple': True,
            'products': product_list,
        })


def _find_column_index(headers, candidates):
    """헤더 리스트에서 후보 컬럼명 중 매칭되는 인덱스를 반환한다."""
    headers_lower = [h.lower().replace(' ', '') for h in headers]
    for candidate in candidates:
        candidate_lower = candidate.lower().replace(' ', '')
        for idx, h in enumerate(headers_lower):
            if candidate_lower in h or h in candidate_lower:
                return idx
    return None


# ============================================================================
# API: 세션 관리
# ============================================================================

@csrf_exempt
@login_required
@staff_required
@require_POST
def create_session(request):
    """새 재고조사 세션을 시작한다."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    name = data.get('name', '').strip()
    if not name:
        return JsonResponse({'error': '세션명을 입력해주세요.'}, status=400)

    # 이미 활성 세션이 있는지 확인
    active = InventorySession.objects.filter(status='active').first()
    if active:
        return JsonResponse({
            'error': f'이미 진행 중인 세션이 있습니다: {active.name}',
        }, status=400)

    session = InventorySession.objects.create(
        name=name,
        started_by=request.user.name if hasattr(request.user, 'name') else str(request.user),
    )

    return JsonResponse({
        'success': True,
        'session': _session_to_dict(session),
    })


@csrf_exempt
@login_required
@staff_required
@require_POST
def end_session(request, session_id):
    """재고조사 세션을 종료한다."""
    try:
        session = InventorySession.objects.get(pk=session_id)
    except InventorySession.DoesNotExist:
        return JsonResponse({'error': '세션을 찾을 수 없습니다.'}, status=404)

    if session.status == 'closed':
        return JsonResponse({'error': '이미 종료된 세션입니다.'}, status=400)

    session.status = 'closed'
    session.ended_at = timezone.now()
    session.save(update_fields=['status', 'ended_at'])

    return JsonResponse({
        'success': True,
        'session': _session_to_dict(session),
    })


@login_required
@staff_required
@require_GET
def get_sessions(request):
    """세션 목록을 조회한다."""
    sessions = InventorySession.objects.annotate(
        record_count=Count('records'),
    ).order_by('-started_at')

    return JsonResponse({
        'sessions': [
            {
                **_session_to_dict(s),
                'record_count': s.record_count,
            }
            for s in sessions
        ],
    })


# ============================================================================
# API: 스캔
# ============================================================================

@csrf_exempt
@login_required
@staff_required
@require_POST
def scan_location(request):
    """로케이션 바코드를 스캔한다. 미등록이면 자동 생성."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    barcode = data.get('barcode', '').strip()
    if not barcode:
        return JsonResponse({'error': '로케이션 바코드를 입력해주세요.'}, status=400)

    # 활성 세션 확인
    active_session = InventorySession.objects.filter(status='active').first()
    if not active_session:
        return JsonResponse({'error': '진행 중인 세션이 없습니다.'}, status=400)

    location, created = Location.objects.get_or_create(
        barcode=barcode,
        defaults={'name': '', 'zone': ''},
    )

    # 현재 세션에서 이 로케이션에 등록된 기록 수
    record_count = InventoryRecord.objects.filter(
        session=active_session,
        location=location,
    ).count()

    return JsonResponse({
        'success': True,
        'created': created,
        'location': {
            'id': location.pk,
            'barcode': location.barcode,
            'name': location.name,
            'zone': location.zone,
        },
        'record_count': record_count,
    })


@csrf_exempt
@login_required
@staff_required
@require_POST
def scan_product(request):
    """상품을 등록한다.

    동일 세션+로케이션+바코드+상품명 조합이 이미 있으면 수량을 합산한다.
    바코드만 입력하고 상품명이 없는 경우:
      - 상품 마스터에 단일 상품이 있으면 자동 보완
      - 없으면 에러 반환 (수기 입력 유도)
    """
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)

    session_id = data.get('session_id')
    location_id = data.get('location_id')
    barcode = data.get('barcode', '').strip()
    product_name = data.get('product_name', '').strip()
    quantity = data.get('quantity', 1)
    expiry_date = data.get('expiry_date', '').strip()
    lot_number = data.get('lot_number', '').strip()

    if not session_id or not location_id:
        return JsonResponse({'error': '세션과 로케이션 정보가 필요합니다.'}, status=400)

    if not barcode and not product_name:
        return JsonResponse({'error': '상품바코드 또는 상품명을 입력해주세요.'}, status=400)

    try:
        quantity = int(quantity)
        if quantity < 1:
            quantity = 1
    except (ValueError, TypeError):
        quantity = 1

    # 세션 확인
    try:
        session = InventorySession.objects.get(pk=session_id, status='active')
    except InventorySession.DoesNotExist:
        return JsonResponse({'error': '활성 세션을 찾을 수 없습니다.'}, status=400)

    # 로케이션 확인
    try:
        location = Location.objects.get(pk=location_id)
    except Location.DoesNotExist:
        return JsonResponse({'error': '로케이션을 찾을 수 없습니다.'}, status=400)

    # 상품 마스터에서 상품명 자동 보완 (단일 상품만)
    if barcode and not product_name:
        master_products = Product.objects.filter(barcode=barcode)
        cnt = master_products.count()
        if cnt == 1:
            product_name = master_products.first().display_name or master_products.first().name
        elif cnt == 0:
            # 미등록 바코드 — 자동 등록 차단
            return JsonResponse({
                'error': f'등록되지 않은 바코드입니다: {barcode}\n상품 관리에서 먼저 등록하거나, 상품명을 수기로 입력해주세요.',
            }, status=400)
        else:
            # 복수 상품 — 클라이언트에서 선택해야 함
            return JsonResponse({
                'error': '동일 바코드에 여러 상품이 있습니다. 상품을 선택해주세요.',
            }, status=400)

    if not product_name:
        return JsonResponse({'error': '상품명을 입력해주세요.'}, status=400)

    # 동일 세션+로케이션+바코드+상품명+유통기한+로트번호이면 수량 합산
    existing = InventoryRecord.objects.filter(
        session=session,
        location=location,
        barcode=barcode,
        product_name=product_name,
        expiry_date=expiry_date,
        lot_number=lot_number,
    ).first()

    worker_name = request.user.name if hasattr(request.user, 'name') else str(request.user)

    if existing:
        existing.quantity += quantity
        existing.worker = worker_name
        existing.save(update_fields=['quantity', 'worker'])
        record = existing
    else:
        record = InventoryRecord.objects.create(
            session=session,
            location=location,
            barcode=barcode,
            product_name=product_name,
            quantity=quantity,
            expiry_date=expiry_date,
            lot_number=lot_number,
            worker=worker_name,
        )

    return JsonResponse({
        'success': True,
        'record': _record_to_dict(record),
    })


# ============================================================================
# API: 기록 조회 / 삭제
# ============================================================================

@login_required
@staff_required
@require_GET
def get_records(request):
    """재고 기록을 조회한다.

    Query params:
        session_id: 세션 ID (필수)
        group_by: 'location' 또는 'product' (기본: location)
        search: 검색어 (로케이션 바코드, 상품명, 상품바코드)
    """
    session_id = request.GET.get('session_id')
    if not session_id:
        return JsonResponse({'error': 'session_id가 필요합니다.'}, status=400)

    group_by = request.GET.get('group_by', 'location')
    search = request.GET.get('search', '').strip()

    records = InventoryRecord.objects.filter(
        session_id=session_id,
    ).select_related('location')

    if search:
        records = records.filter(
            models_Q_search(search)
        )

    if group_by == 'product':
        return _get_records_by_product(records)
    else:
        return _get_records_by_location(records)


@csrf_exempt
@login_required
@staff_required
def delete_record(request, record_id):
    """재고 기록을 삭제한다."""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE 메서드만 허용됩니다.'}, status=405)

    try:
        record = InventoryRecord.objects.get(pk=record_id)
    except InventoryRecord.DoesNotExist:
        return JsonResponse({'error': '기록을 찾을 수 없습니다.'}, status=404)

    record.delete()
    return JsonResponse({'success': True})


@login_required
@staff_required
@require_GET
def get_location_records(request):
    """특정 로케이션의 현재 세션 기록을 조회한다.

    Query params:
        session_id: 세션 ID
        location_id: 로케이션 ID
    """
    session_id = request.GET.get('session_id')
    location_id = request.GET.get('location_id')

    if not session_id or not location_id:
        return JsonResponse({'error': 'session_id와 location_id가 필요합니다.'}, status=400)

    records = InventoryRecord.objects.filter(
        session_id=session_id,
        location_id=location_id,
    ).order_by('-created_at')

    return JsonResponse({
        'records': [_record_to_dict(r) for r in records],
    })


# ============================================================================
# 헬퍼 함수
# ============================================================================

def models_Q_search(search):
    """검색어로 Q 필터를 생성한다."""
    from django.db.models import Q
    return (
        Q(barcode__icontains=search) |
        Q(product_name__icontains=search) |
        Q(location__barcode__icontains=search) |
        Q(location__name__icontains=search)
    )


def _get_records_by_location(records):
    """로케이션별로 그룹핑하여 반환한다."""
    grouped = defaultdict(lambda: {
        'location': None,
        'products': [],
        'total_quantity': 0,
    })

    for r in records:
        key = r.location_id
        if grouped[key]['location'] is None:
            grouped[key]['location'] = {
                'id': r.location.pk,
                'barcode': r.location.barcode,
                'name': r.location.name,
                'zone': r.location.zone,
            }
        grouped[key]['products'].append(_record_to_dict(r))
        grouped[key]['total_quantity'] += r.quantity

    # 로케이션 바코드 순 정렬
    result = sorted(
        grouped.values(),
        key=lambda x: x['location']['barcode'] if x['location'] else '',
    )

    return JsonResponse({'groups': result})


def _get_records_by_product(records):
    """상품별로 그룹핑하여 반환한다."""
    grouped = defaultdict(lambda: {
        'barcode': '',
        'product_name': '',
        'locations': [],
        'total_quantity': 0,
    })

    for r in records:
        key = r.barcode or r.product_name
        if not grouped[key]['barcode']:
            grouped[key]['barcode'] = r.barcode
            grouped[key]['product_name'] = r.product_name
        grouped[key]['locations'].append({
            'location_barcode': r.location.barcode,
            'location_name': r.location.name,
            'quantity': r.quantity,
            'expiry_date': r.expiry_date,
            'lot_number': r.lot_number,
            'record_id': r.pk,
            'worker': r.worker,
            'created_at': timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M'),
        })
        grouped[key]['total_quantity'] += r.quantity

    # 상품 바코드 순 정렬
    result = sorted(
        grouped.values(),
        key=lambda x: x['barcode'] or x['product_name'],
    )

    return JsonResponse({'groups': result})


def _session_to_dict(session):
    """InventorySession 인스턴스를 dict로 변환한다."""
    return {
        'id': session.pk,
        'name': session.name,
        'status': session.status,
        'status_display': session.get_status_display(),
        'started_at': timezone.localtime(session.started_at).strftime('%Y-%m-%d %H:%M'),
        'ended_at': (
            timezone.localtime(session.ended_at).strftime('%Y-%m-%d %H:%M')
            if session.ended_at else None
        ),
        'started_by': session.started_by,
    }


def _record_to_dict(record):
    """InventoryRecord 인스턴스를 dict로 변환한다."""
    return {
        'id': record.pk,
        'barcode': record.barcode,
        'product_name': record.product_name,
        'quantity': record.quantity,
        'expiry_date': record.expiry_date,
        'lot_number': record.lot_number,
        'worker': record.worker,
        'location_barcode': record.location.barcode if hasattr(record, 'location') and record.location else '',
        'location_name': record.location.name if hasattr(record, 'location') and record.location else '',
        'created_at': timezone.localtime(record.created_at).strftime('%Y-%m-%d %H:%M'),
    }
