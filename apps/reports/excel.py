"""
리포트 엑셀 생성 유틸리티
"""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from django.http import HttpResponse


def build_workbook(data, columns, sheet_title='Report'):
    """데이터를 openpyxl Workbook으로 변환

    Args:
        data: list[dict] — 각 dict는 한 행
        columns: list[tuple(key, header_label, width)]
        sheet_title: 시트 이름

    Returns:
        openpyxl.Workbook
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 헤더
    for col_idx, (key, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, (key, _label, _width) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
            cell.border = thin_border

    return wb


def workbook_to_response(wb, filename):
    """Workbook → HttpResponse (Excel 다운로드)"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def workbook_to_file(wb, filename):
    """Workbook → ContentFile (ReportFile.file.save 용)"""
    buf = io.BytesIO()
    wb.save(buf)
    return ContentFile(buf.getvalue(), name=filename)
