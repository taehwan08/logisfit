"""
일일택배사업부 리포트 서비스

사방넷 출고 엑셀 파일을 파싱하여 브랜드별 단포/합포 출고건수를 집계하고,
DB에 저장된 데이터로부터 엑셀 리포트를 생성한다.

pandas 미사용 — openpyxl(.xlsx) + xlrd(.xls) 만 사용.
"""
import os
from collections import defaultdict
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import xlrd


# ============================================================================
# 엑셀 파싱 (업로드 파일 → dict)
# ============================================================================

def _read_rows_xlsx(uploaded_file):
    """xlsx 파일에서 헤더 + 데이터 행 반환"""
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError('엑셀 파일이 비어 있습니다.')
    return rows[0], rows[1:]


def _read_rows_xls(uploaded_file):
    """xls 파일에서 헤더 + 데이터 행 반환"""
    content = uploaded_file.read()
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        raise ValueError('엑셀 파일이 비어 있습니다.')
    header = [ws.cell_value(0, c) for c in range(ws.ncols)]
    data_rows = []
    for r in range(1, ws.nrows):
        data_rows.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
    return header, data_rows


def parse_parcel_excel(uploaded_file):
    """
    사방넷 출고 엑셀을 파싱하여 집계 결과를 dict로 반환.

    Returns:
        {
            'brands': {
                '브랜드A': {'single': 85, 'combo': 35},
                '브랜드B': {'single': 41, 'combo': 0},
                ...
            },
            'total_orders': 161,
            'total_single': 126,
            'total_combo': 35,
        }
    """
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext == '.xls':
        header, data_rows = _read_rows_xls(uploaded_file)
    else:
        header, data_rows = _read_rows_xlsx(uploaded_file)

    num_cols = len(header)
    if num_cols < 66:
        raise ValueError(
            f'컬럼 수가 부족합니다. (필요: 66개 이상, 현재: {num_cols}개)\n'
            '사방넷 출고 데이터 엑셀 파일인지 확인해주세요.'
        )

    brands = defaultdict(lambda: {'single': 0, 'combo': 0})

    for row in data_rows:
        if len(row) < 66:
            continue
        brand = str(row[64]).strip() if row[64] is not None else ''
        memo = str(row[65]).strip() if row[65] is not None else ''
        if not brand:
            continue
        if '합' in memo:
            brands[brand]['combo'] += 1
        else:
            brands[brand]['single'] += 1

    if not brands:
        raise ValueError('집계할 데이터가 없습니다. 파일 내용을 확인해주세요.')

    total_single = sum(v['single'] for v in brands.values())
    total_combo = sum(v['combo'] for v in brands.values())

    return {
        'brands': dict(brands),
        'total_orders': total_single + total_combo,
        'total_single': total_single,
        'total_combo': total_combo,
    }


# ============================================================================
# 엑셀 생성 (DB 데이터 → xlsx)
# ============================================================================

def generate_report_excel(report):
    """
    DailyParcelReport 인스턴스로부터 스타일이 적용된 엑셀 BytesIO를 생성.

    Args:
        report: DailyParcelReport 인스턴스 (.brands prefetch 권장)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '일일택배사업부'

    # 스타일
    h_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='4472C4')
    h_align = Alignment(horizontal='center', vertical='center')
    n_font = Font(name='Arial', size=10)
    b_font = Font(name='Arial', bold=True, size=10)
    c_align = Alignment(horizontal='center', vertical='center')
    l_align = Alignment(horizontal='left', vertical='center')
    r_align = Alignment(horizontal='right', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 헤더
    headers = ['브랜드', '포장구분', '출고건수', '비율(%)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border

    # 데이터
    row_num = 2
    brand_list = report.brands.all().order_by('brand_name')
    for brand in brand_list:
        for i, (pack_type, count) in enumerate([('단포', brand.single_count), ('합포', brand.combo_count)]):
            ws.cell(row=row_num, column=1, value=brand.brand_name if i == 0 else '').font = b_font if i == 0 else n_font
            ws.cell(row=row_num, column=1).alignment = l_align
            ws.cell(row=row_num, column=1).border = border

            ws.cell(row=row_num, column=2, value=pack_type).font = n_font
            ws.cell(row=row_num, column=2).alignment = c_align
            ws.cell(row=row_num, column=2).border = border

            ws.cell(row=row_num, column=3, value=count).font = n_font
            ws.cell(row=row_num, column=3).alignment = r_align
            ws.cell(row=row_num, column=3).number_format = '#,##0'
            ws.cell(row=row_num, column=3).border = border

            ratio = round(count / brand.total_count * 100, 1) if brand.total_count else 0
            ws.cell(row=row_num, column=4, value=ratio).font = n_font
            ws.cell(row=row_num, column=4).alignment = r_align
            ws.cell(row=row_num, column=4).number_format = '0.0'
            ws.cell(row=row_num, column=4).border = border

            row_num += 1

    # 합계 행
    total_fill = PatternFill('solid', fgColor='D9E2F3')
    for ci, val in enumerate([
        '합계', '', report.total_orders,
        f'{report.combo_ratio}',
    ], 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.font = b_font
        cell.fill = total_fill
        cell.alignment = c_align if ci <= 2 else r_align
        cell.border = border
    ws.cell(row=row_num, column=3).number_format = '#,##0'

    # 열 너비
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
